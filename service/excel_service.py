import os
import sys
# 获取项目根目录
current_dir = os.path.dirname(os.path.abspath(__file__))
root_dir = os.path.dirname(current_dir)
# 将项目根目录添加到Python路径
sys.path.append(root_dir)

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional, Union, Set, Tuple
import logging
from service.mysql_service import MySQLService
from configs.db_config import *
from utils.timer import timer
import time

class ExcelService:
    """Excel服务类，用于更新Excel文件中的数据"""
    
    def __init__(self, mysql_service: MySQLService):
        """
        初始化Excel服务
        
        参数:
        mysql_service (MySQLService): MySQL服务实例
        """
        self.mysql_service = mysql_service
        
    def _load_workbook(self, excel_path: str) -> Tuple[str, Any]:
        """
        加载Excel工作簿
        
        参数:
        excel_path (str): Excel文件路径
        
        返回:
        Tuple[str, Workbook]: (新文件路径, 工作簿对象)
        """
        print(f"正在加载Excel文件: {excel_path}")
        start_time = time.time()
        
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel文件不存在: {excel_path}")
        
        # 创建新的文件名（原名+_update）
        file_name, file_ext = os.path.splitext(excel_path)
        new_excel_path = f"{file_name}_update{file_ext}"
        
        # 加载工作簿，保持数据链接
        workbook = load_workbook(excel_path, data_only=False)
        
        print(f"Excel文件加载完成，耗时: {time.time() - start_time:.2f}秒")
        return new_excel_path, workbook
    
    def _collect_formulas(self, worksheet, columns_to_check: Set[str]) -> Dict[str, str]:
        """
        收集工作表中的公式
        
        参数:
        worksheet: 工作表对象
        columns_to_check (Set[str]): 需要检查的列集合
        
        返回:
        Dict[str, str]: 公式字典，格式为 {cell_coord: formula}
        """
        print(f"正在收集公式...")
        start_time = time.time()
        
        all_formulas = {}
        formula_count = 0
        
        # 只检查指定的列
        for col_letter in columns_to_check:
            for row in range(1, worksheet.max_row + 1):
                cell_coord = f"{col_letter}{row}"
                cell = worksheet[cell_coord]
                if cell.data_type == 'f':
                    all_formulas[cell_coord] = cell.value
                    formula_count += 1
        
        print(f"公式收集完成，找到 {formula_count} 个公式，耗时: {time.time() - start_time:.2f}秒")
        return all_formulas
    
    def _clear_old_data(self, worksheet, column_mapping: Dict[str, str], 
                       all_formulas: Dict[str, str]) -> int:
        """
        清除旧数据
        
        参数:
        worksheet: 工作表对象
        column_mapping (Dict[str, str]): 列映射
        all_formulas (Dict[str, str]): 公式字典
        
        返回:
        int: 清除的单元格数量
        """
        print("正在清除旧数据...")
        start_time = time.time()
        
        cleared_cells = 0
        for row in range(2, worksheet.max_row + 1):
            for excel_col in column_mapping.keys():
                cell_coord = f"{excel_col}{row}"
                if cell_coord not in all_formulas:
                    worksheet[cell_coord].value = None
                    cleared_cells += 1
        
        print(f"旧数据清除完成，清除了 {cleared_cells} 个单元格，耗时: {time.time() - start_time:.2f}秒")
        return cleared_cells
    
    def _update_data(self, worksheet, mysql_data: pd.DataFrame, 
                    column_mapping: Dict[str, str], all_formulas: Dict[str, str]) -> Set[int]:
        """
        更新数据
        
        参数:
        worksheet: 工作表对象
        mysql_data (pd.DataFrame): MySQL数据
        column_mapping (Dict[str, str]): 列映射
        all_formulas (Dict[str, str]): 公式字典
        
        返回:
        Set[int]: 更新的行集合
        """
        print(f"正在更新数据，数据行数: {len(mysql_data)}")
        start_time = time.time()
        
        updated_rows = set()
        updated_cells = 0
        
        for idx, mysql_row in mysql_data.iterrows():
            excel_row = idx + 2  # 从第2行开始写入，保留表头
            updated_rows.add(excel_row)
            
            for excel_col, mysql_col in column_mapping.items():
                cell_coord = f"{excel_col}{excel_row}"
                
                # 跳过公式单元格
                if cell_coord in all_formulas:
                    continue
                    
                if mysql_col in mysql_data.columns:
                    value = mysql_row[mysql_col]
                    worksheet[cell_coord].value = value
                    updated_cells += 1
        
        print(f"数据更新完成，更新了 {updated_cells} 个单元格，涉及 {len(updated_rows)} 行，耗时: {time.time() - start_time:.2f}秒")
        return updated_rows
    
    def _restore_formulas(self, worksheet, all_formulas: Dict[str, str], 
                         updated_rows: Set[int]) -> int:
        """
        恢复公式
        
        参数:
        worksheet: 工作表对象
        all_formulas (Dict[str, str]): 公式字典
        updated_rows (Set[int]): 更新的行集合
        
        返回:
        int: 恢复的公式数量
        """
        print("正在恢复公式...")
        start_time = time.time()
        
        restored_count = 0
        for coord, formula in all_formulas.items():
            # 提取行号
            col_letter = ''.join(filter(str.isalpha, coord))
            row_num = int(''.join(filter(str.isdigit, coord)))
            
            # 只恢复表头行和有数据的行的公式
            if row_num == 1 or row_num in updated_rows:
                worksheet[coord].value = formula
                restored_count += 1
        
        print(f"公式恢复完成，恢复了 {restored_count} 个公式，耗时: {time.time() - start_time:.2f}秒")
        return restored_count
    
    def _save_workbook(self, workbook, new_excel_path: str) -> None:
        """
        保存工作簿
        
        参数:
        workbook: 工作表对象
        new_excel_path (str): 新文件路径
        """
        print(f"正在保存文件: {new_excel_path}")
        start_time = time.time()
        
        workbook.save(new_excel_path)
        
        print(f"文件保存完成，耗时: {time.time() - start_time:.2f}秒")
    
    @timer
    def update_excel_file(
        self,
        excel_path: str,
        mysql_data: pd.DataFrame,
        sheet_updates: Dict[str, Dict[str, Any]]
    ) -> str:
        """
        使用传入的MySQL数据更新Excel文件中的特定表格和行
        
        参数:
        excel_path (str): Excel文件路径
        mysql_data (pd.DataFrame): 从MySQL查询得到的数据框，包含需要更新到Excel的数据
        sheet_updates (Dict[str, Dict[str, Any]]): 需要更新的表格和相关信息的字典
            格式: {
                "sheet_name": {
                    "table": "mysql_table_name",
                    "column_mapping": {
                        "A": "mysql_col1",  # Excel列号: MySQL列名
                        "B": "mysql_col2",
                        "C": "mysql_col3"
                    },
                    "known_formula_columns": ["D", "E"]  # 已知包含公式的列（可选）
                }
            }
        
        返回:
        str: 更新后的Excel文件路径
        """
        try:
            # 1. 加载工作簿
            new_excel_path, workbook = self._load_workbook(excel_path)
            
            # 2. 处理每个工作表
            for sheet_name, update_info in sheet_updates.items():
                if sheet_name not in workbook.sheetnames:
                    logging.warning(f"表格 {sheet_name} 不存在于Excel文件中")
                    continue
                
                print(f"\n开始处理工作表: {sheet_name}")
                
                # 获取工作表
                worksheet = workbook[sheet_name]
                
                if mysql_data.empty:
                    logging.warning(f"MySQL数据为空")
                    continue
                
                # 获取列映射和已知公式列
                column_mapping = update_info.get("column_mapping", {})
                known_formula_columns = set(update_info.get("known_formula_columns", []))
                
                # 确定需要检查的列
                columns_to_check = set(column_mapping.keys())
                if known_formula_columns:
                    columns_to_check.update(known_formula_columns)
                
                # 3. 收集公式
                all_formulas = self._collect_formulas(worksheet, columns_to_check)
                
                # 4. 清除旧数据
                self._clear_old_data(worksheet, column_mapping, all_formulas)
                
                # 5. 更新数据
                updated_rows = self._update_data(worksheet, mysql_data, column_mapping, all_formulas)
                
                # 6. 恢复公式
                self._restore_formulas(worksheet, all_formulas, updated_rows)
            
            # 7. 保存工作簿
            self._save_workbook(workbook, new_excel_path)
            
            return new_excel_path
            
        except Exception as e:
            logging.error(f"更新Excel文件时出错: {e}")
            raise
        
    @timer
    def update_excel_file_multi_sources(
        self,
        excel_path: str,
        data_sources: Dict[str, pd.DataFrame],
        sheet_updates: Dict[str, List[Dict[str, Any]]]
    ) -> str:
        """
        使用多个数据源一次性更新Excel文件中的多个表格
        
        参数:
        excel_path (str): Excel文件路径
        data_sources (Dict[str, pd.DataFrame]): 数据源字典，键为数据源名称，值为DataFrame
        sheet_updates (Dict[str, List[Dict[str, Any]]]): 更新配置
            格式: {
                "sheet_name": [
                    {
                        "data_source": "source_name",  # 对应data_sources中的键
                        "column_mapping": {
                            "A": "col1",  # Excel列: DataFrame列
                            "B": "col2"
                        },
                        "known_formula_columns": ["C", "D"]  # 可选，已知包含公式的列
                    },
                    {
                        "data_source": "another_source",
                        "column_mapping": {
                            "E": "col1",
                            "F": "col2"
                        }
                    }
                ]
            }
        
        返回:
        str: 更新后的Excel文件路径
        """
        try:
            # 1. 加载工作簿
            new_excel_path, workbook = self._load_workbook(excel_path)
            
            # 2. 处理每个工作表
            for sheet_name, updates_list in sheet_updates.items():
                if sheet_name not in workbook.sheetnames:
                    logging.warning(f"表格 {sheet_name} 不存在于Excel文件中")
                    continue
                
                print(f"\n开始处理工作表: {sheet_name}")
                
                # 获取工作表
                worksheet = workbook[sheet_name]
                
                # 获取所有需要检查的列
                all_columns = set()
                known_formula_columns = set()
                
                for update in updates_list:
                    all_columns.update(update.get("column_mapping", {}).keys())
                    if "known_formula_columns" in update:
                        known_formula_columns.update(update["known_formula_columns"])
                
                # 确定需要检查的列
                columns_to_check = all_columns.union(known_formula_columns)
                
                # 3. 收集公式
                all_formulas = self._collect_formulas(worksheet, columns_to_check)
                
                # 记录所有更新的行
                all_updated_rows = set()
                
                # 4. 处理每个更新配置
                for update_info in updates_list:
                    data_source_name = update_info.get("data_source")
                    if data_source_name not in data_sources:
                        logging.warning(f"数据源 {data_source_name} 不存在")
                        continue
                    
                    mysql_data = data_sources[data_source_name]
                    if mysql_data.empty:
                        logging.warning(f"数据源 {data_source_name} 为空")
                        continue
                    
                    column_mapping = update_info.get("column_mapping", {})
                    
                    # 5. 清除旧数据
                    self._clear_old_data(worksheet, column_mapping, all_formulas)
                    
                    # 6. 更新数据
                    updated_rows = self._update_data(worksheet, mysql_data, column_mapping, all_formulas)
                    all_updated_rows.update(updated_rows)
                
                # 7. 恢复公式
                self._restore_formulas(worksheet, all_formulas, all_updated_rows)
            
            # 8. 保存工作簿
            self._save_workbook(workbook, new_excel_path)
            
            return new_excel_path
            
        except Exception as e:
            logging.error(f"更新Excel文件时出错: {e}")
            raise
        
    async def update_excel_file_async(self, excel_path: str, sheet_updates: Dict[str, List[Dict]]):
        try:
            # 创建线程池执行数据库查询
            with ThreadPoolExecutor() as executor:
                # 并行查询所有需要的数据
                tasks = []
                for sheet_name, tables_info in sheet_updates.items():
                    for table_info in tables_info:
                        task = self.loop.run_in_executor(
                            executor,
                            self.mysql_service.get_monthly_yearly_data,
                            table_info["table"],
                            table_info.get("columns", ["*"]),
                            table_info.get("time_column", "日期"),
                            table_info.get("current_month", True)
                        )
                        tasks.append((table_info, task))
                
                # 等待所有查询完成
                results = []
                for table_info, task in tasks:
                    mysql_data = await task
                    results.append((table_info, mysql_data))
                
                # 更新Excel文件
                workbook = load_workbook(excel_path, data_only=False)
                for table_info, mysql_data in results:
                    self._update_sheet_data(
                        worksheet=workbook[sheet_name],
                        mysql_data=mysql_data,
                        column_mapping=table_info["column_mapping"]
                    )
                
                workbook.save(excel_path)
                
        except Exception as e:
            logging.error(f"异步更新Excel文件时出错: {e}")
            raise
            
    def read_excel_sheet(
        self,
        excel_path: str,
        sheet_name: str,
        usecols: Optional[List[str]] = None
    ) -> pd.DataFrame:
        """
        读取Excel表格中的特定sheet
        
        参数:
        excel_path (str): Excel文件路径
        sheet_name (str): 表格名称
        usecols (List[str], optional): 需要读取的列名列表
        
        返回:
        pd.DataFrame: 包含表格数据的DataFrame
        """
        try:
            df = pd.read_excel(
                excel_path,
                sheet_name=sheet_name,
                usecols=usecols
            )
            return df
        except Exception as e:
            logging.error(f"读取Excel表格时出错: {e}")
            raise
        
async def main():
    excel_service = ExcelService(mysql_service)
    await excel_service.update_excel_file_async(
        excel_path="/home/yangsongze/szdn/2025-01试结算市场数据※ - 模版 copy.xlsx",
        sheet_updates=sheet_updates
    )
    
if __name__ == "__main__":
    # 初始化服务
    mysql_service = MySQLService(host=MYSQL_IP, user=MYSQL_ACCOUNT, 
                               password=MYSQL_PASSWORD, database=MYSQL_DBNAME)
    excel_service = ExcelService(mysql_service)
    excel_path = "/home/yangsongze/szdn/2025-01试结算市场数据※ - 模版 copy.xlsx"

    # 1. 获取预测数据并更新
    forecast_data = mysql_service.get_monthly_yearly_data(
        table="load_forecast_one_day_in_advance",
        columns=["日期", "时刻", "统调负荷", "省内B类电源", "西电东送电力", 
                 "省内A类电源", "地方电源出力", "粤港联络线"],
        time_column="日期",
        current_month=True
    )
    
    sheet_updates_forecast = {
        "统调负荷": {
            "table": "load_forecast_one_day_in_advance",
            "column_mapping": {
                "C": "日期", 
                "D": "时刻", 
                "E": "统调负荷", 
                "F": "省内B类电源", 
                "G": "西电东送电力", 
                "H": "省内A类电源",
                "I": "地方电源出力",
                "J": "粤港联络线"
            }
        }
    }
    excel_service.update_excel_file(excel_path, forecast_data, sheet_updates_forecast)

    # # 2. 获取实际运行数据并更新
    # actual_data = mysql_service.get_monthly_yearly_data(
    #     table="load_actual_running",
    #     columns=["日期", "时刻", "统调负荷", "省内B类电源", "西电东送电力", 
    #              "省内A类电源", "地方电源出力", "粤港联络线"],
    #     time_column="日期",
    #     current_month=True
    # )
    
    # sheet_updates_actual = {
    #     "统调负荷": {
    #         "table": "load_actual_running",
    #         "column_mapping": {
    #             "M": "时刻", 
    #             "N": "统调负荷", 
    #             "O": "省内B类电源", 
    #             "P": "西电东送电力", 
    #             "Q": "省内A类电源",
    #             "R": "地方电源出力",
    #             "S": "粤港联络线"
    #         }
    #     }
    # }
    # excel_service.update_excel_file(excel_path, actual_data, sheet_updates_actual)

    # # 3. 获取实时节点价格并更新
    # real_time_price = mysql_service.get_monthly_yearly_data(
    #     table="average_real_time_node_price",
    #     columns=["日期", "时刻", "价格"],
    #     time_column="日期",
    #     current_month=True
    # )
    
    # sheet_updates_real_time = {
    #     "统调负荷": {
    #         "table": "average_real_time_node_price",
    #         "column_mapping": {
    #             "AC": "价格"
    #         }
    #     }
    # }
    # excel_service.update_excel_file(excel_path, real_time_price, sheet_updates_real_time)

    # # 4. 获取日前节点价格并更新
    # day_ahead_price = mysql_service.get_monthly_yearly_data(
    #     table="average_day_ahead_node_price",
    #     columns=["日期", "时刻", "价格"],
    #     time_column="日期",
    #     current_month=True
    # )
    
    # sheet_updates_day_ahead = {
    #     "统调负荷": {
    #         "table": "average_day_ahead_node_price",
    #         "column_mapping": {
    #             "AB": "价格"
    #         }
    #     }
    # }
    # excel_service.update_excel_file(excel_path, day_ahead_price, sheet_updates_day_ahead)

    print("所有数据更新完成！")


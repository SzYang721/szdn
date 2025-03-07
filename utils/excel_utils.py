import string
from typing import Dict, List, Any, Optional, Tuple, Union
import pandas as pd
from openpyxl.utils import get_column_letter as openpyxl_get_column_letter
from openpyxl.utils import column_index_from_string

def get_column_letter(col_idx: int) -> str:
    """
    将列索引转换为Excel列字母
    
    参数:
    col_idx (int): 列索引（从1开始）
    
    返回:
    str: Excel列字母
    
    示例:
    get_column_letter(1)  # 返回 'A'
    get_column_letter(27)  # 返回 'AA'
    """
    return openpyxl_get_column_letter(col_idx)

def get_column_index(col_letter: str) -> int:
    """
    将Excel列字母转换为列索引
    
    参数:
    col_letter (str): Excel列字母
    
    返回:
    int: 列索引（从1开始）
    
    示例:
    get_column_index('A')  # 返回 1
    get_column_index('AA')  # 返回 27
    """
    return column_index_from_string(col_letter)

def excel_range(start_cell: str, end_cell: str) -> List[str]:
    """
    生成Excel单元格范围
    
    参数:
    start_cell (str): 起始单元格，如 'A1'
    end_cell (str): 结束单元格，如 'C3'
    
    返回:
    List[str]: 单元格列表
    
    示例:
    excel_range('A1', 'B2')  # 返回 ['A1', 'A2', 'B1', 'B2']
    """
    # 解析起始和结束单元格
    start_col = ''.join(filter(str.isalpha, start_cell))
    start_row = int(''.join(filter(str.isdigit, start_cell)))
    
    end_col = ''.join(filter(str.isalpha, end_cell))
    end_row = int(''.join(filter(str.isdigit, end_cell)))
    
    # 获取列索引
    start_col_idx = get_column_index(start_col)
    end_col_idx = get_column_index(end_col)
    
    # 生成单元格列表
    cells = []
    for col_idx in range(start_col_idx, end_col_idx + 1):
        col_letter = get_column_letter(col_idx)
        for row in range(start_row, end_row + 1):
            cells.append(f"{col_letter}{row}")
    
    return cells

def dataframe_to_excel(df: pd.DataFrame, excel_path: str, sheet_name: str = 'Sheet1',
                      start_cell: str = 'A1', index: bool = False, 
                      header: bool = True) -> None:
    """
    将DataFrame写入Excel文件
    
    参数:
    df (pd.DataFrame): 要写入的DataFrame
    excel_path (str): Excel文件路径
    sheet_name (str): 工作表名称
    start_cell (str): 起始单元格
    index (bool): 是否写入索引
    header (bool): 是否写入表头
    """
    # 解析起始单元格
    start_col = ''.join(filter(str.isalpha, start_cell))
    start_row = int(''.join(filter(str.isdigit, start_cell)))
    
    # 计算起始行和列
    start_row_idx = start_row - 1  # pandas使用0索引
    start_col_idx = get_column_index(start_col) - 1  # pandas使用0索引
    
    # 写入Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name=sheet_name, startrow=start_row_idx, 
                   startcol=start_col_idx, index=index, header=header)

def find_formulas_in_excel(excel_path: str, sheet_name: Optional[str] = None) -> Dict[str, Dict[str, str]]:
    """
    查找Excel文件中的所有公式
    
    参数:
    excel_path (str): Excel文件路径
    sheet_name (str, optional): 工作表名称，如果为None则查找所有工作表
    
    返回:
    Dict[str, Dict[str, str]]: 公式字典，格式为 {sheet_name: {cell_coord: formula}}
    """
    from openpyxl import load_workbook
    
    workbook = load_workbook(excel_path, data_only=False)
    
    formulas = {}
    
    if sheet_name is not None:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")
        
        sheets_to_check = [workbook[sheet_name]]
    else:
        sheets_to_check = workbook.worksheets
    
    for sheet in sheets_to_check:
        sheet_formulas = {}
        
        for row in sheet.rows:
            for cell in row:
                if cell.data_type == 'f':
                    sheet_formulas[cell.coordinate] = cell.value
        
        if sheet_formulas:
            formulas[sheet.title] = sheet_formulas
    
    return formulas

def extract_cell_info(cell_coord: str) -> Tuple[str, int]:
    """
    从单元格坐标提取列字母和行号
    
    参数:
    cell_coord (str): 单元格坐标，如 'A1'
    
    返回:
    Tuple[str, int]: (列字母, 行号)
    
    示例:
    extract_cell_info('A1')  # 返回 ('A', 1)
    extract_cell_info('AB123')  # 返回 ('AB', 123)
    """
    col_letter = ''.join(filter(str.isalpha, cell_coord))
    row_num = int(''.join(filter(str.isdigit, cell_coord)))
    
    return col_letter, row_num 